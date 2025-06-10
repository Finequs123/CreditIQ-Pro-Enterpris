import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any

def validate_csv_columns(df: pd.DataFrame) -> Dict[str, Any]:
    """Validate CSV column structure"""
    required_columns = [
        'pan', 'age', 'monthly_income', 'credit_score', 'foir', 'dpd30plus',
        'enquiry_count', 'credit_vintage', 'loan_mix_type', 'loan_completion_ratio',
        'defaulted_loans', 'unsecured_loan_amount', 'outstanding_amount_percent',
        'our_lender_exposure', 'channel_type', 'writeoff_flag'
    ]
    
    errors = []
    missing_columns = []
    
    # Check for missing columns
    for col in required_columns:
        if col not in df.columns:
            missing_columns.append(col)
    
    if missing_columns:
        errors.append(f"Missing required columns: {', '.join(missing_columns)}")
    
    # Check data types and ranges for existing columns
    if 'age' in df.columns:
        invalid_ages = df[(df['age'] < 18) | (df['age'] > 80)]
        if not invalid_ages.empty:
            errors.append(f"Invalid age values found in rows: {list(invalid_ages.index + 1)}")
    
    if 'monthly_income' in df.columns:
        invalid_income = df[df['monthly_income'] < 0]
        if not invalid_income.empty:
            errors.append(f"Negative income values found in rows: {list(invalid_income.index + 1)}")
    
    if 'credit_score' in df.columns:
        invalid_credit = df[(df['credit_score'] < -1) | (df['credit_score'] > 900)]
        if not invalid_credit.empty:
            errors.append(f"Invalid credit score values found in rows: {list(invalid_credit.index + 1)}")
    
    if 'foir' in df.columns:
        invalid_foir = df[(df['foir'] < 0) | (df['foir'] > 2)]
        if not invalid_foir.empty:
            errors.append(f"Invalid FOIR values found in rows: {list(invalid_foir.index + 1)}")
    
    # Check for valid loan mix types
    if 'loan_mix_type' in df.columns:
        valid_loan_types = ["PL/HL/CC", "Gold + Consumer Durable", "Only Gold", "Agri/Other loans"]
        invalid_loan_types = df[~df['loan_mix_type'].isin(valid_loan_types)]
        if not invalid_loan_types.empty:
            errors.append(f"Invalid loan mix types found in rows: {list(invalid_loan_types.index + 1)}")
    
    # Check for valid channel types
    if 'channel_type' in df.columns:
        valid_channels = ["Merchant/Referral", "Digital/Other"]
        invalid_channels = df[~df['channel_type'].isin(valid_channels)]
        if not invalid_channels.empty:
            errors.append(f"Invalid channel types found in rows: {list(invalid_channels.index + 1)}")
    
    return {
        'valid': len(errors) == 0,
        'errors': errors
    }

def create_excel_output(applicant_data_list: List[Dict[str, Any]], 
                       results_list: List[Dict[str, Any]], 
                       is_bulk: bool = False) -> bytes:
    """Create Excel output for scoring results"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Summary data
    summary_data = []
    total_applications = len(results_list)
    
    if total_applications > 0:
        bucket_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0}
        total_score = 0
        
        for result in results_list:
            bucket = result.get('final_bucket', 'D')
            bucket_counts[bucket] += 1
            total_score += result.get('final_score', 0)
        
        avg_score = total_score / total_applications
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Applications', total_applications],
            ['Average Score', f"{avg_score:.2f}"],
            ['Auto-Approve (A)', bucket_counts['A']],
            ['Recommend (B)', bucket_counts['B']],
            ['Refer (C)', bucket_counts['C']],
            ['Decline (D)', bucket_counts['D']],
            ['Approval Rate', f"{((bucket_counts['A'] + bucket_counts['B']) / total_applications * 100):.1f}%"]
        ]
    else:
        summary_data = [
            ['Metric', 'Value'],
            ['Total Applications', 0]
        ]
    
    # Write summary data
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
    
    # Auto-adjust column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create detailed results sheet
    details_ws = wb.create_sheet("Detailed Results")
    
    # Prepare detailed data
    detailed_data = []
    headers = [
        'PAN', 'Age', 'Monthly Income', 'Credit Score', 'FOIR', 'DPD30+', 'Enquiry Count',
        'Credit Vintage', 'Loan Mix Type', 'Completion Ratio', 'Defaulted Loans',
        'Unsecured Amount', 'Outstanding %', 'Our Exposure', 'Channel Type',
        'Final Score', 'Initial Bucket', 'Final Bucket', 'Decision', 'Bucket Movements'
    ]
    detailed_data.append(headers)
    
    for i, (applicant, result) in enumerate(zip(applicant_data_list, results_list)):
        movements = ""
        if result.get('bucket_movements'):
            movements = "; ".join([f"{m['from']}→{m['to']}" for m in result['bucket_movements']])
        
        row = [
            applicant.get('pan', ''),
            applicant.get('age', ''),
            applicant.get('monthly_income', ''),
            applicant.get('credit_score', ''),
            applicant.get('foir', ''),
            applicant.get('dpd30plus', ''),
            applicant.get('enquiry_count', ''),
            applicant.get('credit_vintage', ''),
            applicant.get('loan_mix_type', ''),
            applicant.get('loan_completion_ratio', ''),
            applicant.get('defaulted_loans', ''),
            applicant.get('unsecured_loan_amount', ''),
            applicant.get('outstanding_amount_percent', ''),
            applicant.get('our_lender_exposure', ''),
            applicant.get('channel_type', ''),
            f"{result.get('final_score', 0):.2f}",
            result.get('initial_bucket', ''),
            result.get('final_bucket', ''),
            result.get('decision', ''),
            movements
        ]
        detailed_data.append(row)
    
    # Write detailed data
    for row_idx, row_data in enumerate(detailed_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = details_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
    
    # Auto-adjust column widths for details sheet
    for column in details_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        details_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create variable scores sheet
    if results_list and results_list[0].get('variable_scores'):
        scores_ws = wb.create_sheet("Variable Scores")
        
        # Prepare variable scores data
        variable_headers = ['PAN'] + list(results_list[0]['variable_scores'].keys())
        scores_data = [variable_headers]
        
        for applicant, result in zip(applicant_data_list, results_list):
            if result.get('variable_scores'):
                row = [applicant.get('pan', '')]
                for var_name in variable_headers[1:]:
                    var_score = result['variable_scores'].get(var_name, {})
                    row.append(f"{var_score.get('weighted_score', 0):.2f}")
                scores_data.append(row)
        
        # Write variable scores data
        for row_idx, row_data in enumerate(scores_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = scores_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
        
        # Auto-adjust column widths for scores sheet
        for column in scores_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            scores_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.read()

def format_currency(amount: float) -> str:
    """Format currency values"""
    if amount >= 10000000:  # 1 crore
        return f"₹{amount/10000000:.1f}Cr"
    elif amount >= 100000:  # 1 lakh
        return f"₹{amount/100000:.1f}L"
    elif amount >= 1000:  # 1 thousand
        return f"₹{amount/1000:.1f}K"
    else:
        return f"₹{amount:,.0f}"

def format_percentage(value: float) -> str:
    """Format percentage values"""
    return f"{value*100:.1f}%"
